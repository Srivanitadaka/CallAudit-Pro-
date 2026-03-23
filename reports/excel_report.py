# reports/excel_report.py
"""
Excel Report Generator
─────────────────────────────────────────────
Exports all scored calls to a formatted Excel file.

Usage:
  from reports.excel_report import generate_excel
  excel_bytes = generate_excel(results_dir)
"""

import io
import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── Colours ────────────────────────────────────────────
BG_DARK    = "080C14"
BG_SURFACE = "0D1320"
BG_HEAD    = "111827"
BG_GREEN   = "166534"
BG_YELLOW  = "713F12"
BG_RED     = "7F1D1D"
BG_ORANGE  = "7C2D12"
BG_BLUE    = "0C4A6E"

FG_WHITE   = "E2E8F0"
FG_MUTED   = "64748B"
FG_GREEN   = "22C55E"
FG_YELLOW  = "F59E0B"
FG_RED     = "F87171"
FG_ORANGE  = "FB923C"
FG_ACCENT  = "38BDF8"

GRADE_BG = {"A": BG_GREEN, "B": BG_BLUE, "C": BG_YELLOW, "D": BG_ORANGE, "F": BG_RED}
GRADE_FG = {"A": FG_GREEN, "B": FG_ACCENT, "C": FG_YELLOW, "D": FG_ORANGE, "F": FG_RED}


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color=FG_WHITE, bold=False, size=11):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")

def _align(horizontal="left", vertical="center", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def _border():
    s = Side(style="thin", color="1E293B")
    return Border(left=s, right=s, top=s, bottom=s)

def _score_color(score, mx=100):
    pct = score / mx
    if pct >= 0.75: return FG_GREEN
    if pct >= 0.55: return FG_YELLOW
    return FG_RED

def _score_bg(score, mx=100):
    pct = score / mx
    if pct >= 0.75: return BG_GREEN
    if pct >= 0.55: return BG_YELLOW
    return BG_RED


# ══════════════════════════════════════════════════════
# MAIN FUNCTION
# ══════════════════════════════════════════════════════
def generate_excel(results_dir: str = "analysis_results") -> bytes:
    """
    Generate Excel report from all scored_*.json files.
    Returns Excel file as bytes.
    """
    results_path = Path(results_dir)
    files        = sorted(results_path.glob("scored_*.json"))

    if not files:
        raise ValueError("No scored results found in analysis_results/")

    # Load all results
    all_results = []
    for f in files:
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            data["_filename"] = f.name
            all_results.append(data)
        except Exception:
            pass

    wb = Workbook()

    # ── Build all sheets ───────────────────────────────
    _sheet_summary(wb, all_results)
    _sheet_all_calls(wb, all_results)
    _sheet_violations(wb, all_results)
    _sheet_improvements(wb, all_results)
    _sheet_agent_quality(wb, all_results)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ══════════════════════════════════════════════════════
# SHEET 1 — SUMMARY OVERVIEW
# ══════════════════════════════════════════════════════
def _sheet_summary(wb, results):
    ws = wb.create_sheet("📊 Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.tab_color = "38BDF8"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "CallAudit Pro — Quality Report"
    ws["A1"].font      = _font(FG_ACCENT, bold=True, size=16)
    ws["A1"].fill      = _fill(BG_DARK)
    ws["A1"].alignment = _align("center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws["A2"].font      = _font(FG_MUTED, size=10)
    ws["A2"].fill      = _fill(BG_DARK)
    ws["A2"].alignment = _align("center")

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 20

    # KPI cards row
    scores    = [r.get("overall_score", 0) for r in results]
    avg_score = sum(scores) / len(scores) if scores else 0
    pass_rate = sum(1 for s in scores if s >= 60) / len(scores) * 100 if scores else 0
    total_v   = sum(len(r.get("violations", [])) for r in results)
    resolved  = sum(1 for r in results if r.get("was_resolved", False))

    grade_counts = {}
    for r in results:
        g = r.get("grade", "?")
        grade_counts[g] = grade_counts.get(g, 0) + 1

    kpis = [
        ("Total Calls",    len(results),         FG_WHITE,  BG_SURFACE),
        ("Avg Score",      f"{avg_score:.1f}/100",FG_ACCENT, BG_BLUE),
        ("Pass Rate ≥60",  f"{pass_rate:.0f}%",  FG_GREEN,  BG_GREEN),
        ("Total Violations",int(total_v),         FG_RED,    BG_RED),
        ("Resolved",       int(resolved),         FG_GREEN,  BG_GREEN),
        ("Grade F Calls",  grade_counts.get("F",0),FG_RED,  BG_RED),
    ]

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 16

    for col, (label, val, fg, bg) in enumerate(kpis, 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 18

        # Value cell
        cell_val = ws.cell(row=5, column=col, value=val)
        cell_val.font      = _font(fg, bold=True, size=18)
        cell_val.fill      = _fill(bg)
        cell_val.alignment = _align("center")

        # Label cell
        cell_lbl = ws.cell(row=6, column=col, value=label)
        cell_lbl.font      = _font(FG_MUTED, size=9)
        cell_lbl.fill      = _fill(BG_SURFACE)
        cell_lbl.alignment = _align("center")

    # Grade breakdown
    ws.row_dimensions[9].height = 20
    ws["A9"] = "Grade Distribution"
    ws["A9"].font = _font(FG_WHITE, bold=True, size=12)
    ws["A9"].fill = _fill(BG_DARK)

    headers = ["Grade", "Count", "Percentage", "Avg Score"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=10, column=col, value=h)
        c.font      = _font(FG_MUTED, bold=True, size=10)
        c.fill      = _fill(BG_HEAD)
        c.alignment = _align("center")
        c.border    = _border()

    grade_results = {}
    for r in results:
        g = r.get("grade", "?")
        if g not in grade_results:
            grade_results[g] = []
        grade_results[g].append(r.get("overall_score", 0))

    row = 11
    for grade in ["A", "B", "C", "D", "F"]:
        if grade not in grade_results:
            continue
        gscores  = grade_results[grade]
        count    = len(gscores)
        pct      = count / len(results) * 100
        avg      = sum(gscores) / count

        fg = GRADE_FG.get(grade, FG_WHITE)
        bg = GRADE_BG.get(grade, BG_SURFACE)

        data = [f"Grade {grade}", count, f"{pct:.1f}%", f"{avg:.1f}/100"]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(bg if col == 1 else BG_SURFACE)
            c.font      = _font(fg if col == 1 else FG_WHITE, bold=(col==1))
            c.alignment = _align("center")
            c.border    = _border()
        row += 1

    # Dimension averages
    ws.row_dimensions[row+1].height = 20
    ws.cell(row=row+1, column=1, value="Dimension Score Averages").font = _font(FG_WHITE, bold=True, size=12)
    ws.cell(row=row+1, column=1).fill = _fill(BG_DARK)

    dim_headers = ["Dimension", "Average Score", "Out Of", "Rating"]
    for col, h in enumerate(dim_headers, 1):
        c = ws.cell(row=row+2, column=col, value=h)
        c.font      = _font(FG_MUTED, bold=True, size=10)
        c.fill      = _fill(BG_HEAD)
        c.alignment = _align("center")
        c.border    = _border()

    DIM_LABELS = {
        "empathy":                  "Empathy",
        "professionalism":          "Professionalism",
        "compliance":               "Compliance",
        "resolution_effectiveness": "Resolution",
        "communication_clarity":    "Clarity",
    }
    dim_row = row + 3
    for key, label in DIM_LABELS.items():
        vals = [
            r.get("dimension_scores", r.get("scores", {})).get(key, 0)
            for r in results
        ]
        avg = sum(vals) / len(vals) if vals else 0
        rating = "Excellent" if avg>=8 else "Good" if avg>=6 else "Fair" if avg>=4 else "Poor"
        fg = FG_GREEN if avg>=8 else FG_YELLOW if avg>=6 else FG_ORANGE if avg>=4 else FG_RED

        row_data = [label, f"{avg:.1f}", "10", rating]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=dim_row, column=col, value=val)
            c.fill      = _fill(BG_SURFACE)
            c.font      = _font(fg if col in [2,4] else FG_WHITE)
            c.alignment = _align("center")
            c.border    = _border()
        dim_row += 1


# ══════════════════════════════════════════════════════
# SHEET 2 — ALL CALLS
# ══════════════════════════════════════════════════════
def _sheet_all_calls(wb, results):
    ws = wb.create_sheet("📋 All Calls")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "22C55E"
    ws.freeze_panes = "A2"

    headers = [
        "File", "Grade", "Score", "Sentiment", "Outcome",
        "Resolved", "Violations", "Sat Rating",
        "Empathy", "Prof", "Compliance", "Resolution", "Clarity",
        "Issue Detected", "Summary"
    ]
    col_widths = [30, 10, 10, 14, 18, 12, 12, 12, 10, 10, 12, 12, 10, 40, 60]

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _font(FG_ACCENT, bold=True, size=10)
        c.fill      = _fill(BG_HEAD)
        c.alignment = _align("center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24

    # Data rows
    for row_idx, r in enumerate(results, 2):
        grade = r.get("grade", "?")
        score = r.get("overall_score", 0)
        dims  = r.get("dimension_scores", r.get("scores", {}))
        sat   = r.get("satisfaction", {})

        row_data = [
            r.get("_filename", "").replace("scored_","").replace(".json",""),
            f"Grade {grade}",
            score,
            r.get("sentiment", "neutral").title(),
            r.get("call_outcome", "Unknown"),
            "✔ Yes" if r.get("was_resolved") else "✘ No",
            len(r.get("violations", [])),
            sat.get("rating", 0),
            dims.get("empathy", 0),
            dims.get("professionalism", 0),
            dims.get("compliance", 0),
            dims.get("resolution_effectiveness", 0),
            dims.get("communication_clarity", 0),
            str(r.get("issue_detected", ""))[:100],
            str(r.get("summary", ""))[:200],
        ]

        bg = BG_SURFACE if row_idx % 2 == 0 else BG_DARK

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = _fill(bg)
            c.alignment = _align("center" if col <= 13 else "left", wrap=(col >= 14))
            c.border    = _border()

            # Color specific columns
            if col == 2:  # Grade
                c.font = _font(GRADE_FG.get(grade, FG_WHITE), bold=True)
                c.fill = _fill(GRADE_BG.get(grade, BG_SURFACE))
            elif col == 3:  # Score
                c.font = _font(_score_color(score))
            elif col == 6:  # Resolved
                c.font = _font(FG_GREEN if r.get("was_resolved") else FG_RED)
            elif col == 7:  # Violations
                vcount = len(r.get("violations", []))
                c.font = _font(FG_RED if vcount > 0 else FG_GREEN)
            elif col in [9,10,11,12,13]:  # Dimension scores
                c.font = _font(_score_color(val or 0, 10))
            else:
                c.font = _font(FG_WHITE)

        ws.row_dimensions[row_idx].height = 20


# ══════════════════════════════════════════════════════
# SHEET 3 — VIOLATIONS
# ══════════════════════════════════════════════════════
def _sheet_violations(wb, results):
    ws = wb.create_sheet("⚠ Violations")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "F87171"
    ws.freeze_panes = "A2"

    headers = ["File", "Grade", "Score", "Severity", "Type", "Quote", "Explanation"]
    col_widths = [28, 10, 10, 12, 28, 50, 60]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _font(FG_RED, bold=True, size=10)
        c.fill      = _fill(BG_HEAD)
        c.alignment = _align("center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24

    SEV_FG = {"critical":FG_RED,"high":FG_ORANGE,"medium":FG_YELLOW,"low":FG_MUTED}
    SEV_BG = {"critical":BG_RED,"high":BG_ORANGE,"medium":BG_YELLOW,"low":BG_SURFACE}

    row_idx = 2
    for r in results:
        grade = r.get("grade","?")
        score = r.get("overall_score",0)
        fname = r.get("_filename","").replace("scored_","").replace(".json","")

        for v in r.get("violations", []):
            sev = (v.get("severity") or "medium").lower()
            row_data = [
                fname,
                f"Grade {grade}",
                score,
                sev.upper(),
                (v.get("type") or "").replace("_"," ").title(),
                str(v.get("quote",""))[:200],
                str(v.get("explanation",""))[:300],
            ]
            bg = BG_SURFACE if row_idx % 2 == 0 else BG_DARK

            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border    = _border()
                c.alignment = _align("center" if col<=4 else "left", wrap=(col>=5))

                if col == 2:
                    c.font = _font(GRADE_FG.get(grade,FG_WHITE), bold=True)
                    c.fill = _fill(GRADE_BG.get(grade,BG_SURFACE))
                elif col == 3:
                    c.font = _font(_score_color(score))
                    c.fill = _fill(bg)
                elif col == 4:
                    c.font = _font(SEV_FG.get(sev,FG_WHITE), bold=True)
                    c.fill = _fill(SEV_BG.get(sev,BG_SURFACE))
                else:
                    c.font = _font(FG_WHITE)
                    c.fill = _fill(bg)

            ws.row_dimensions[row_idx].height = 32
            row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No violations found ✓").font = _font(FG_GREEN, bold=True)


# ══════════════════════════════════════════════════════
# SHEET 4 — IMPROVEMENTS
# ══════════════════════════════════════════════════════
def _sheet_improvements(wb, results):
    ws = wb.create_sheet("💡 Improvements")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "38BDF8"
    ws.freeze_panes = "A2"

    headers = ["File", "Grade", "Score", "Area", "Suggestion", "Example Phrase"]
    col_widths = [28, 10, 10, 22, 60, 60]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _font(FG_ACCENT, bold=True, size=10)
        c.fill      = _fill(BG_HEAD)
        c.alignment = _align("center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24
    row_idx = 2

    for r in results:
        grade = r.get("grade","?")
        score = r.get("overall_score",0)
        fname = r.get("_filename","").replace("scored_","").replace(".json","")

        for i in r.get("improvements", []):
            row_data = [
                fname,
                f"Grade {grade}",
                score,
                (i.get("area") or "").replace("_"," ").title(),
                str(i.get("suggestion",""))[:300],
                str(i.get("example",""))[:200],
            ]
            bg = BG_SURFACE if row_idx % 2 == 0 else BG_DARK

            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border    = _border()
                c.alignment = _align("center" if col<=3 else "left", wrap=(col>=4))

                if col == 2:
                    c.font = _font(GRADE_FG.get(grade,FG_WHITE), bold=True)
                    c.fill = _fill(GRADE_BG.get(grade,BG_SURFACE))
                elif col == 3:
                    c.font = _font(_score_color(score))
                    c.fill = _fill(bg)
                elif col == 4:
                    c.font = _font(FG_ACCENT)
                    c.fill = _fill(bg)
                else:
                    c.font = _font(FG_WHITE)
                    c.fill = _fill(bg)

            ws.row_dimensions[row_idx].height = 32
            row_idx += 1


# ══════════════════════════════════════════════════════
# SHEET 5 — AGENT QUALITY
# ══════════════════════════════════════════════════════
def _sheet_agent_quality(wb, results):
    ws = wb.create_sheet("👔 Agent Quality")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "F59E0B"
    ws.freeze_panes = "A2"

    headers = [
        "File", "Grade", "Score",
        "Lang Clarity", "Professionalism", "Time Efficiency",
        "Response Eff", "Empathy Score",
        "Bias Detected", "Customer Calmed", "Sentiment", "Frustration"
    ]
    col_widths = [28, 10, 10, 14, 16, 16, 14, 14, 14, 16, 14, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _font(FG_YELLOW, bold=True, size=10)
        c.fill      = _fill(BG_HEAD)
        c.alignment = _align("center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24

    for row_idx, r in enumerate(results, 2):
        grade = r.get("grade","?")
        score = r.get("overall_score",0)
        aq    = r.get("agent_quality",{})
        sat   = r.get("satisfaction",{})
        fname = r.get("_filename","").replace("scored_","").replace(".json","")

        row_data = [
            fname,
            f"Grade {grade}",
            score,
            aq.get("language_clarity",0),
            aq.get("professionalism",0),
            aq.get("time_efficiency",0),
            aq.get("response_efficiency",0),
            aq.get("empathy_score",0),
            "⚠ Yes" if aq.get("bias_detected") else "✔ No",
            "✔ Yes" if aq.get("calmed_customer") else "✘ No",
            str(sat.get("sentiment","")).title(),
            str(sat.get("customer_frustration","None")),
        ]

        bg = BG_SURFACE if row_idx % 2 == 0 else BG_DARK

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = _fill(bg)
            c.alignment = _align("center")
            c.border    = _border()

            if col == 2:
                c.font = _font(GRADE_FG.get(grade,FG_WHITE), bold=True)
                c.fill = _fill(GRADE_BG.get(grade,BG_SURFACE))
            elif col == 3:
                c.font = _font(_score_color(score))
            elif col in [4,5,6,7]:
                c.font = _font(_score_color(val or 0, 20))
            elif col == 8:
                c.font = _font(_score_color(val or 0, 10))
            elif col == 9:
                c.font = _font(FG_RED if aq.get("bias_detected") else FG_GREEN)
            elif col == 10:
                c.font = _font(FG_GREEN if aq.get("calmed_customer") else FG_RED)
            else:
                c.font = _font(FG_WHITE)

        ws.row_dimensions[row_idx].height = 20